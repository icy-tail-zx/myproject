#! /usr/bin/env python
# coding=utf-8

import re
import time
import datetime
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains

def get_excel():
    option = webdriver.ChromeOptions()
    option.add_argument('headless')
    path = r'chromedriver.exe'
    browser = webdriver.Chrome(options=option, executable_path=path)
    # browser = webdriver.Chrome(executable_path=path)
    get_url = 'http://pbi.vinda.com/'
    browser.get(get_url)
    time.sleep(3)
    # browser.maximize_window()
    browser.set_window_size(1920, 1080)
    # 登录
    browser.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div[1]/input').send_keys('BIDEV')
    browser.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div[2]/input').send_keys('bidev')
    browser.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div[3]/button').click()
    time.sleep(3)
    browser.find_element_by_xpath('/html/body/div[1]/div[1]/div/ul/li[2]/div').click()
    browser.execute_script('window.scrollTo(0,document.body.scrollHeight)')     # 翻滚到底
    time.sleep(2)
    # 日期
    this_year = time.strftime("%Y", time.localtime(time.time()))
    this_day = time.strftime("%Y/%m/%d", time.localtime(time.time()))
    today_rhz = this_day.replace('/0', '/')

    # Sheet---总览
    def sheet_zl():
        browser.find_element_by_xpath('/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[1]/div[1]/span[2]/span/span').click()
        time.sleep(30)      # 等待数据加载30s
        browser.switch_to.frame('refreshAlarm')
        time.sleep(1)
        # 今日时间
        jinrishijian = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[90]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][1]/*[name()="text"]/*[name()="tspan"]').text
        jinrishijian = jinrishijian.replace('-', '/')
        # 今日销售额
        jinrixiaoshou = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div').text
        # 本月销售额
        benyuexiaoshou = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[4]/transform/div/div[3]/div/visual-modern/div').text
        # 本月目标额
        benyuemubiao = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[41]/transform/div/div[3]/div').text
        # 总达成
        zongdacheng = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[72]/transform/div/div[3]/div/visual-modern/div').text
        # 时间已过
        shijianyiguo = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[71]/transform/div/div[3]/div/visual-modern/div').text.split('过')[1].strip("）")
        # 上周日均
        shangzhourijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[45]/transform/div/div[3]/div/visual-modern/div').text
        # 本周累计销售
        benzhouleijixiaoshou = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[44]/transform/div/div[3]/div/visual-modern/div').text
        # 上周累计销售
        shangzhouleijixiaoshou = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[45]/transform/div/div[3]/div/visual-modern/div').text
        # 上月同期累计
        shangyuetongqileiji = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[82]/transform/div/div[3]/div/visual-modern/div').text
        # 缺口
        quekou = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[69]/transform/div/div[3]/div').text
        # 浏览量
        liulanliang = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[17]/transform/div/div[3]/div/visual-modern/div').text
        liulanliang_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[18]/transform/div/div[3]/div/visual-modern/div').text
        liulanliang_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[46]/transform/div/div[3]/div/visual-modern/div').text
        # 总订单数
        zongdingdanshu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[35]/transform/div/div[3]/div/visual-modern/div').text
        zongdingdanshu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[36]/transform/div/div[3]/div/visual-modern/div').text
        zongdingdanshu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[47]/transform/div/div[3]/div/visual-modern/div').text
        # 总访客数
        zongfangkeshu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[13]/transform/div/div[3]/div/visual-modern/div').text
        zongfangkeshu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[14]/transform/div/div[3]/div/visual-modern/div').text
        zongfangkeshu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[55]/transform/div/div[3]/div/visual-modern/div').text
        # 老访客数
        laofangkeshu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[31]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laofangkeshu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', laofangkeshu.replace(',', '')).group()
        laofangkeshu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[15]/transform/div/div[3]/div/visual-modern/div').text
        laofangkeshu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[57]/transform/div/div[3]/div/visual-modern/div').text
        # 新访客数
        xinfangkeshu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[31]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinfangkeshu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', xinfangkeshu.replace(',', '')).group()
        xinfangkeshu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[16]/transform/div/div[3]/div/visual-modern/div').text
        xinfangkeshu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[56]/transform/div/div[3]/div/visual-modern/div').text
        # 总买家数
        zongmaijiashu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[19]/transform/div/div[3]/div/visual-modern/div').text
        zongmaijiashu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[20]/transform/div/div[3]/div/visual-modern/div').text
        zongmaijiashu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[58]/transform/div/div[3]/div/visual-modern/div').text
        # 老买家数
        laomaijiashu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[32]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laomaijiashu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', laomaijiashu.replace(',', '')).group()
        laomaijiashu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[21]/transform/div/div[3]/div/visual-modern/div').text
        laomaijiashu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[60]/transform/div/div[3]/div/visual-modern/div').text
        # 新买家数
        xinmaijiashu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[32]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinmaijiashu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', xinmaijiashu.replace(',', '')).group()
        xinmaijiashu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[22]/transform/div/div[3]/div/visual-modern/div').text
        xinmaijiashu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[59]/transform/div/div[3]/div/visual-modern/div').text
        # 下单率
        xiadanlv = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[33]/transform/div/div[3]/div/visual-modern/div').text
        xiadanlv_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[34]/transform/div/div[3]/div/visual-modern/div').text
        # 老访客下单率
        laoxiadanlv = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[39]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laoxiadanlv_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[51]/transform/div/div[3]/div/visual-modern/div').text
        laoxiadanlv_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[52]/transform/div/div[3]/div/visual-modern/div').text
        # 新访客下单率
        xinxiadanlv = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[39]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinxiadanlv_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[53]/transform/div/div[3]/div/visual-modern/div').text
        xinxiadanlv_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[54]/transform/div/div[3]/div/visual-modern/div').text
        # 客单价
        kedanjia = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[23]/transform/div/div[3]/div/visual-modern/div').text
        kedanjia_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[24]/transform/div/div[3]/div/visual-modern/div').text
        kedanjia_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[91]/transform/div/div[3]/div/visual-modern/div').text
        # 老客单价
        laokedanjia = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[40]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laokedanjia_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[37]/transform/div/div[3]/div/visual-modern/div').text
        laokedanjia_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[62]/transform/div/div[3]/div/visual-modern/div').text
        # 新客单价
        xinkedanjia = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[40]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinkedanjia_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[38]/transform/div/div[3]/div/visual-modern/div').text
        xinkedanjia_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[61]/transform/div/div[3]/div/visual-modern/div').text
        # 主动评价数
        zhudongpingjiashu = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[25]/transform/div/div[3]/div/visual-modern/div').text
        zhudongpingjiashu_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[26]/transform/div/div[3]/div/visual-modern/div').text
        zhudongpingjiashu_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[65]/transform/div/div[3]/div/visual-modern/div').text
        # 正面评价
        zhengmianpingjia = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[30]/transform/div/div[3]/div/visual-modern/div').text
        zhengmianpingjia_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[29]/transform/div/div[3]/div/visual-modern/div').text
        zhengmianpingjia_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[64]/transform/div/div[3]/div/visual-modern/div').text
        # 负面评价
        fumianpingjia = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[28]/transform/div/div[3]/div/visual-modern/div').text
        fumianpingjia_rijun = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[27]/transform/div/div[3]/div/visual-modern/div').text
        fumianpingjia_hb = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[63]/transform/div/div[3]/div/visual-modern/div').text
        # 淘宝群数据
        data_dict_tb = {
            '活跃人数': [],
            '今日引导成交金额': [],
            '进店购买转化率': [],
            '进店浏览转化率': [],
            '引导成交人数': [],
            '引导进店人数': [],
            '引导客单价': [],
        }
        data_all_tb = browser.find_element_by_class_name('bodyCells')       # Xpath路径太长
        i = 1
        for data_name_tb, data_list_tb in data_dict_tb.items():
            tb1 = data_all_tb.find_element_by_xpath(f'./div/div/div[2]/div[{i}]').text
            tb2 = data_all_tb.find_element_by_xpath(f'./div/div/div[3]/div[{i}]').text
            tb3 = data_all_tb.find_element_by_xpath(f'./div/div/div[4]/div[{i}]').text
            tb4 = data_all_tb.find_element_by_xpath(f'./div/div/div[5]/div[{i}]').text
            data_list_tb.append(tb1)
            data_list_tb.append(tb2)
            data_list_tb.append(tb3)
            data_list_tb.append(tb4)
            i = i + 1
        # 打开Excel模板
        wb = load_workbook('Libresse天猫旗舰店日报表.xlsx')
        ws1 = wb['总览']
        # 数据写入'汇总'
        benyuedijitian = int(jinrishijian.split('/')[2])
        ws1['A1'] = jinrishijian    # 界面拿数据
        ws1['C2'] = benyuexiaoshou
        ws1['C3'] = jinrixiaoshou
        ws1['E2'] = benyuemubiao
        ws1['E3'] = round(int(benyuexiaoshou.replace(',', ''))/benyuedijitian)
        ws1['G2'] = zongdacheng
        ws1['I2'] = shijianyiguo
        ws1['I3'] = quekou

        ws1['D4'] = benzhouleijixiaoshou
        ws1['F4'] = shangzhouleijixiaoshou
        ws1['H4'] = shangyuetongqileiji
        ws1['D5'] = liulanliang
        ws1['D6'] = zongdingdanshu
        ws1['D7'] = zongfangkeshu
        ws1['D8'] = laofangkeshu
        ws1['D9'] = xinfangkeshu
        ws1['D10'] = zongmaijiashu
        ws1['D11'] = laomaijiashu
        ws1['D12'] = xinmaijiashu
        ws1['D13'] = xiadanlv
        ws1['D14'] = laoxiadanlv
        ws1['D15'] = xinxiadanlv
        ws1['D16'] = kedanjia
        ws1['D17'] = laokedanjia
        ws1['D18'] = xinkedanjia
        ws1['D19'] = zhudongpingjiashu
        ws1['D20'] = zhengmianpingjia
        ws1['D21'] = fumianpingjia

        ws1['H5'] = liulanliang_rijun
        ws1['H6'] = zongdingdanshu_rijun
        ws1['H7'] = zongfangkeshu_rijun
        ws1['H8'] = laofangkeshu_rijun
        ws1['H9'] = xinfangkeshu_rijun
        ws1['H10'] = zongmaijiashu_rijun
        ws1['H11'] = laomaijiashu_rijun
        ws1['H12'] = xinmaijiashu_rijun
        ws1['H13'] = xiadanlv_rijun
        ws1['H14'] = laoxiadanlv_rijun
        ws1['H15'] = xinxiadanlv_rijun
        ws1['H16'] = kedanjia_rijun
        ws1['H17'] = laokedanjia_rijun
        ws1['H18'] = xinkedanjia_rijun
        ws1['H19'] = zhudongpingjiashu_rijun
        ws1['H20'] = zhengmianpingjia_rijun
        ws1['H21'] = fumianpingjia_rijun

        ws1['J5'] = liulanliang_hb
        ws1['J6'] = zongdingdanshu_hb
        ws1['J7'] = zongfangkeshu_hb
        ws1['J8'] = laofangkeshu_hb
        ws1['J9'] = xinfangkeshu_hb
        ws1['J10'] = zongmaijiashu_hb
        ws1['J11'] = laomaijiashu_hb
        ws1['J12'] = xinmaijiashu_hb
        ws1['J14'] = laoxiadanlv_hb
        ws1['J15'] = xinxiadanlv_hb
        ws1['J16'] = kedanjia_hb
        ws1['J17'] = laokedanjia_hb
        ws1['J18'] = xinkedanjia_hb
        ws1['J19'] = zhudongpingjiashu_hb
        ws1['J20'] = zhengmianpingjia_hb
        ws1['J21'] = fumianpingjia_hb

        i = 0
        for data_name_tb, data_list_tb in data_dict_tb.items():
            ws1[f'D{i+22}'] = data_list_tb[0]
            ws1[f'F{i+22}'] = data_list_tb[1]
            ws1[f'H{i+22}'] = data_list_tb[2]
            ws1[f'J{i+22}'] = data_list_tb[3]
            i = i + 1
        browser.switch_to.default_content()  # 切回到原始界面
        wb.save('Libresse天猫旗舰店日报表.xlsx')
        wb.close()

    # Sheet--月汇总
    def sheet_yhz():
        browser.find_element_by_xpath('/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[3]/div[1]/span[2]/span/span').click()
        time.sleep(40)      # 等待数据加载40s
        browser.switch_to.frame('refreshAlarm')
        browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[17]/transform/div/div[3]/div/visual-modern/div/button').click()
        time.sleep(5)
        data_dict_yhz1 = {
            'A21': [],
            'M21': [],
            'F21': [],
            'J21': [],
            'D20': [],
            'N20': [],
            'O20': [],
            'S20': [],
        }
        i = 1
        for data_name_yhz1, data_list_yhz1 in data_dict_yhz1.items():
            data_all_yhz1 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i+2}]/div')
            for data_num in data_all_yhz1:
                a = data_num.text
                data_dict_yhz1[data_name_yhz1].append(a)
            data_all_yhz1 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[3]/div[{i+2}]/div')
            for data_num in data_all_yhz1:
                a = data_num.text
                data_dict_yhz1[data_name_yhz1].append(a)
            i = i + 1
        data_dict_yhz2 = {
            'A20': [],
            'J201': [],
            'J202': [],
            'M20': [],
        }
        i = 1
        for data_name_yhz2, data_list_yhz2 in data_dict_yhz2.items():
            data_all_yhz2 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[2]/div[{i}]/div')
            for data_num in data_all_yhz2:
                b = data_num.text
                data_dict_yhz2[data_name_yhz2].append(b)
            data_all_yhz2 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[4]/div[{i}]/div')
            for data_num in data_all_yhz2:
                b = data_num.text
                data_dict_yhz2[data_name_yhz2].append(b)
            i = i + 1
        # 数据写入月汇总
        wb = load_workbook('Libresse天猫旗舰店日报表.xlsx')
        ws3 = wb['月汇总']
        ws3['A1'] = this_year + '年'
        data_dict_yhz1.update(data_dict_yhz2)
        i = 0
        for data_name_rhz, data_list_rhz in data_dict_yhz1.items():
            for j in range(len(data_list_rhz)):
                ws3.cell(j + 3, i + 2, data_list_rhz[j])
            i = i + 1
        browser.switch_to.default_content()  # 切回到原始界面
        wb.save('Libresse天猫旗舰店日报表.xlsx')

    # Sheet--日汇总2
    def sheet_rhz2():
        # 日期
        riqi_list_1 = []
        # 详细数据_1
        data_dict_1 = {
            '备注': [],
            '今日支付金额': [],
            '老买家支付金额': [],
            '新买家支付金额': [],
            '无线端支付金额': [],
            'PC端支付金额': [],
            '浏览量': [],
            '父订单': [],
            '总访客数': [],
            '老访客数': [],
        }
        # 详细数据_2
        data_dict_2 = {
            '新访客数': [],
            '总买家数': [],
            '老买家数': [],
            '新买家数': [],
            '支付转化率': [],
            '老客户支付转化率': [],
            '新客户支付转化率': [],
            '客单价': [],
            '老客客单价': [],
            '新客客单价': [],
        }
        # 详细数据_3
        data_dict_3 = {
            '入会数': [],
            '入会率': [],
            '累计入会率': [],
            '前台粉丝数': [],
            '后台粉丝数': [],
            '北欧小v巾搜索人气': [],
            'Libresse搜索人气': [],
            '薇尔搜索人气': [],
            '主动评价数': [],
            '正面评价': [],
        }
        # 详细数据_4
        data_dict_4 = {
            '负面评价': [],
            '负评率': [],
            '淘客点击数': [],
            '淘客投入': [],
            '淘客引入金额': [],
            '淘客产出占比': []
        }
        today_rhz = datetime.date.today()
        oneday = datetime.timedelta(days=1)
        start_day_list = str(today_rhz).split('-')
        yesterday_list = str(today_rhz - oneday * 10).split('-')
        # 注意，无头模式部署在服务器上时，日期格式有变化，可用截图观察
        start_day_rhz = start_day_list[1] + '/' + start_day_list[2] + '/' + start_day_list[0]
        yesterday_rhz = yesterday_list[1] + '/' + yesterday_list[2] + '/' + yesterday_list[0]
        browser.find_element_by_xpath('/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[2]/div[1]/span[2]/span/span').click()
        time.sleep(40)  # 等待数据加载40s
        browser.switch_to.frame('refreshAlarm')
        time.sleep(1)
        # 点击输入日期
        def data_input(input_element, text):
            browser.execute_script('arguments[0].removeAttribute(\"readonly\")', input_element)
            input_element.click()
            input_element.click()
            input_element.send_keys(Keys.CONTROL, 'a')
            input_element.send_keys(Keys.DELETE)
            input_element.send_keys(text)
            time.sleep(2)

        # 获取近10天数据
        start_date = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div/div[1]/div/div[1]/input')
        data_input(start_date, yesterday_rhz)
        end_date = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div/div[1]/div/div[2]/input')
        data_input(end_date, start_day_rhz)
        browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/div').click()
        time.sleep(5)
        browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[1]/div').click()
        time.sleep(3)
        # 日期
        riqi_all = browser.find_elements_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[3]/div/div')
        for riqi_1 in riqi_all[0: len(riqi_all) - 1]:
            riqi_list_1.append(riqi_1.text)
        # 详细数据_1
        i = 1
        for data_name, data_list in data_dict_1.items():
            data_all_1 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i}]/div')
            for data_num in data_all_1[0: len(data_all_1) - 1]:
                b = data_num.text
                if b == ' ':
                    b = None
                    data_dict_1[data_name].append(b)
                else:
                    data_dict_1[data_name].append(b)
            i = i + 1
        time.sleep(1)
        btn = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[3]/div[3]')  # 定位滑块
        ActionChains(browser).click_and_hold(on_element=btn).perform()  # 摁住滑块不动
        # ActionChains(browser).move_by_offset(292, 0).perform()
        ActionChains(browser).move_by_offset(570, 0).perform()
        ActionChains(browser).release(on_element=btn).perform()
        time.sleep(2)
        # 详细数据_2
        i = 1
        for data_name, data_list in data_dict_2.items():
            data_all_2 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i}]/div')
            for data_num in data_all_2[0: len(data_all_2) - 1]:
                a = data_num.text
                if a == ' ':
                    a = None
                    data_dict_2[data_name].append(a)
                else:
                    data_dict_2[data_name].append(a)
            i = i + 1
        time.sleep(1)
        # 横向滚动
        btn = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[3]/div[3]')  # 定位滑块
        ActionChains(browser).click_and_hold(on_element=btn).perform()  # 摁住滑块不动
        ActionChains(browser).move_by_offset(570, 0).perform()
        ActionChains(browser).release(on_element=btn).perform()
        time.sleep(2)
        # 详细数据_3
        i = 1
        for data_name, data_list in data_dict_3.items():
            data_all_3 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i}]/div')
            for data_num in data_all_3[0: len(data_all_3) - 1]:
                a = data_num.text
                if a == ' ':
                    a = None
                    data_dict_3[data_name].append(a)
                else:
                    data_dict_3[data_name].append(a)
            i = i + 1
        time.sleep(1)
        # 详细数据_4
        i = 1
        for data_name, data_list in data_dict_4.items():
            data_all_4 = browser.find_elements_by_xpath(f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[2]/div[{i}]/div')
            for data_num in data_all_4[0: len(data_all_4) - 1]:
                a = data_num.text
                if a == ' ':
                    a = None
                    data_dict_4[data_name].append(a)
                else:
                    data_dict_4[data_name].append(a)
            i = i + 1
        btn = browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[3]/div[3]')  # 定位滑块
        ActionChains(browser).click_and_hold(on_element=btn).perform()  # 摁住滑块不动
        ActionChains(browser).move_by_offset(-1140, 0).perform()
        ActionChains(browser).release(on_element=btn).perform()
        time.sleep(3)
        # 合并字典
        data_dict_1.update(data_dict_2)
        data_dict_1.update(data_dict_3)
        data_dict_1.update(data_dict_4)
        # 写入数据
        wb = load_workbook('Libresse天猫旗舰店日报表.xlsx')
        ws2 = wb['日汇总']
        riqi_excel_list = []
        for riqi_excel in ws2['A']:
            riqi_excel_list.append(riqi_excel.value)
        i = 0
        for riqi in riqi_list_1:
            if riqi not in riqi_excel_list:
                ws2.insert_rows(2, 1)
                j = 0
                for dict_name_1, dict_list_1 in data_dict_1.items():
                    ws2.cell(2, 1, riqi)
                    ws2.cell(2, j + 2, dict_list_1[i])
                    j = j + 1
            i = i + 1
        browser.switch_to.default_content()  # 切回到原始界面
        wb.save('Libresse天猫旗舰店日报表.xlsx')

    sheet_rhz2()
    sheet_yhz()
    sheet_zl()
    browser.quit()


