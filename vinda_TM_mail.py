#! /usr/bin/env python
# coding=utf-8

import os
import re
import time
import csv
import smtplib
import datetime
from selenium import webdriver
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By


# 获取Excel数据
def get_excel():
    option = webdriver.ChromeOptions()
    option.add_argument('headless')
    path = r'chromedriver.exe'
    browser = webdriver.Chrome(options=option, executable_path=path)
    browser.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {
        'cmd': 'Page.setDownloadBehavior',
        'params': {
            'behavior': 'allow', 'downloadPath': r"C:\Python"}}
    browser.execute("send_command", params=params)
    get_url = 'http://pbi.vinda.com/'
    browser.get(get_url)
    time.sleep(3)
    browser.set_window_size(1920, 1080)
    # 登录
    browser.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div[1]/input').send_keys('BIDEV')
    browser.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div[2]/input').send_keys('VindaBidev2021')
    browser.find_element_by_xpath('/html/body/div[1]/div/div[2]/div[2]/div[3]/button').click()
    wait = WebDriverWait(browser, 10, 0.5)
    wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/div/ul/li[2]/div')))  # 注意括号
    time.sleep(1)
    browser.find_element_by_xpath('/html/body/div[1]/div[1]/div/ul/li[2]/div').click()
    browser.execute_script('window.scrollTo(0,document.body.scrollHeight)')     # 翻滚到底
    time.sleep(2)
    # 日期
    this_year = time.strftime("%Y", time.localtime(time.time()))
    this_day = time.strftime("%Y/%m/%d", time.localtime(time.time()))
    today_rhz = this_day.replace('/0', '/')

    # Sheet--每日单品汇总
    def sheet_rdp():
        # 获取csv文件
        download_path = r'C:\Python'
        # download_path = r'C:\Users\zx303\Downloads'
        file_from = os.path.join(download_path, 'data.csv')
        file_to = r'C:\Python\Libresse天猫旗舰店日报表.xlsx'
        # file_to = r'C:\Users\zx303\Desktop\Libresse日报表.xlsx'
        if 'data.csv' in os.listdir(download_path):
            os.remove(file_from)
        browser.find_element_by_xpath('/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[4]/div[1]/span[2]/span').click()
        time.sleep(10)  # 等待加载10s
        browser.switch_to.frame('refreshAlarm')  # 再切到内嵌iframe
        time.sleep(1)
        WebDriverWait(browser, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[4]/div[3]')))
        time.sleep(30)
        browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[4]/div[3]').click()
        time.sleep(1)
        browser.find_element_by_xpath('//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/visual-container-header-modern/div/div[1]/div/visual-container-options-menu/visual-header-item-container/div/button').click()
        time.sleep(1)
        browser.find_element_by_xpath('/html/body/div[8]/drop-down-list/ng-transclude/ng-repeat[1]/drop-down-list-item/ng-transclude/ng-switch/div').click()
        time.sleep(2)
        browser.find_element_by_xpath('//*[@id="formatSelect"]').click()
        time.sleep(1)
        browser.find_element_by_xpath('//*[@id="formatSelect"]/option[2]').click()
        time.sleep(2)
        browser.find_element_by_xpath('/html/body/div[8]/div/div/div/div[2]/export-data-dialog/dialog-frame/div/div[2]/section/dialog-footer/button[1]').click()
        time.sleep(30)
        # 数据写入Excel
        # df1 = pd.read_csv(file_from, encoding='utf-8')
        # df2 = df1.sort_values(by='V码')
        # df2.to_csv(file_from, encoding='utf-8')
        wb2 = load_workbook(file_to)
        sheet2 = wb2['每日单品汇总']
        sheet2.delete_rows(2, 200)
        with open(file_from, newline='', encoding='utf-8') as f:
            rows = csv.reader(f)
            a = 1
            for i in rows:
                b = 1
                for j in i:
                    if j == '-':
                        j = None
                    sheet2.cell(a, b, j)
                    b = b + 1
                a = a + 1
        browser.switch_to.default_content()
        wb2.save(file_to)
        wb2.close()

    # Sheet--日汇总
    def sheet_rhz():
        # 日期
        riqi_list_1 = []
        # 详细数据_1
        data_dict_1 = {
            '备注': [],
            '今日支付金额': [],
            '老买家支付金额': [],
            '新买家支付金额': [],
            '无线端支付金额': [],
            'PC端支付金额': [],
            '浏览量': [],
            '父订单': [],
            '总访客数': [],
            '老访客数': [],
        }
        # 详细数据_2
        data_dict_2 = {
            '新访客数': [],
            '总买家数': [],
            '老买家数': [],
            '新买家数': [],
            '支付转化率': [],
            '老客户支付转化率': [],
            '新客户支付转化率': [],
            '客单价': [],
            '老客客单价': [],
            '新客客单价': [],
        }
        # 详细数据_3
        data_dict_3 = {
            '入会数': [],
            '入会率': [],
            '累计入会率': [],
            '前台粉丝数': [],
            '后台粉丝数': [],
            '北欧小v巾搜索人气': [],
            'Libresse搜索人气': [],
            '薇尔搜索人气': [],
            '主动评价数': [],
            '正面评价': [],
        }
        # 详细数据_4
        data_dict_4 = {
            '负面评价': [],
            '负评率': [],
            '淘客点击数': [],
            '淘客投入': [],
            '淘客引入金额': [],
            '淘客产出占比': []
        }
        today_rhz = datetime.date.today()
        oneday = datetime.timedelta(days=1)
        start_day_list = str(today_rhz).split('-')
        yesterday_list = str(today_rhz - oneday * 10).split('-')
        # 注意，无头模式部署在服务器上时，日期格式有变化，可用截图观察
        start_day_rhz = start_day_list[1] + '/' + start_day_list[2] + '/' + start_day_list[0]
        yesterday_rhz = yesterday_list[1] + '/' + yesterday_list[2] + '/' + yesterday_list[0]
        browser.find_element_by_xpath(
            '/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[2]/div[1]/span[2]/span/span').click()
        time.sleep(10)  # 等待数据加载10s
        browser.switch_to.frame('refreshAlarm')
        time.sleep(1)

        # 点击输入日期
        def data_input(input_element, text):
            browser.execute_script('arguments[0].removeAttribute(\"readonly\")', input_element)
            input_element.click()
            input_element.click()
            input_element.send_keys(Keys.CONTROL, 'a')
            input_element.send_keys(Keys.DELETE)
            input_element.send_keys(text)
            time.sleep(2)

        # 获取近10天数据
        WebDriverWait(browser, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div/div[1]/div/div[1]/input')))
        time.sleep(30)
        start_date = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div/div[1]/div/div[1]/input')
        data_input(start_date, yesterday_rhz)
        end_date = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div/div[1]/div/div[2]/input')
        data_input(end_date, start_day_rhz)
        browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[3]/transform/div/div[3]/div/div').click()
        time.sleep(5)
        browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[1]/div').click()
        # time.sleep(3)
        # 日期
        WebDriverWait(browser, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[3]/div/div')))
        time.sleep(30)
        riqi_all = browser.find_elements_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[3]/div/div')
        for riqi_1 in riqi_all[0: len(riqi_all) - 1]:
            riqi_list_1.append(riqi_1.text)
        # 详细数据_1
        i = 1
        for data_name, data_list in data_dict_1.items():
            data_all_1 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i}]/div')
            for data_num in data_all_1[0: len(data_all_1) - 1]:
                b = data_num.text
                if b == ' ':
                    b = None
                    data_dict_1[data_name].append(b)
                else:
                    data_dict_1[data_name].append(b)
            i = i + 1
        time.sleep(1)
        btn = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[3]/div[3]')  # 定位滑块
        ActionChains(browser).click_and_hold(on_element=btn).perform()  # 摁住滑块不动
        # ActionChains(browser).move_by_offset(292, 0).perform()
        ActionChains(browser).move_by_offset(570, 0).perform()
        ActionChains(browser).release(on_element=btn).perform()
        time.sleep(2)
        # 详细数据_2
        i = 1
        for data_name, data_list in data_dict_2.items():
            data_all_2 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i}]/div')
            for data_num in data_all_2[0: len(data_all_2) - 1]:
                a = data_num.text
                if a == ' ':
                    a = None
                    data_dict_2[data_name].append(a)
                else:
                    data_dict_2[data_name].append(a)
            i = i + 1
        time.sleep(1)
        # 横向滚动
        btn = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[3]/div[3]')  # 定位滑块
        ActionChains(browser).click_and_hold(on_element=btn).perform()  # 摁住滑块不动
        ActionChains(browser).move_by_offset(570, 0).perform()
        ActionChains(browser).release(on_element=btn).perform()
        time.sleep(2)
        # 详细数据_3
        i = 1
        for data_name, data_list in data_dict_3.items():
            data_all_3 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i}]/div')
            for data_num in data_all_3[0: len(data_all_3) - 1]:
                a = data_num.text
                if a == ' ':
                    a = None
                    data_dict_3[data_name].append(a)
                else:
                    data_dict_3[data_name].append(a)
            i = i + 1
        time.sleep(1)
        # 详细数据_4
        i = 1
        for data_name, data_list in data_dict_4.items():
            data_all_4 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[2]/div[{i}]/div')
            for data_num in data_all_4[0: len(data_all_4) - 1]:
                a = data_num.text
                if a == ' ':
                    a = None
                    data_dict_4[data_name].append(a)
                else:
                    data_dict_4[data_name].append(a)
            i = i + 1
        btn = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[3]/div[3]')  # 定位滑块
        ActionChains(browser).click_and_hold(on_element=btn).perform()  # 摁住滑块不动
        ActionChains(browser).move_by_offset(-1140, 0).perform()
        ActionChains(browser).release(on_element=btn).perform()
        time.sleep(3)
        # 合并字典
        data_dict_1.update(data_dict_2)
        data_dict_1.update(data_dict_3)
        data_dict_1.update(data_dict_4)
        # 写入数据
        file_to = r'C:\Python\Libresse天猫旗舰店日报表.xlsx'
        # wb = load_workbook('Libresse天猫旗舰店日报表.xlsx')
        wb = load_workbook(file_to)
        ws2 = wb['日汇总']
        riqi_excel_list = []
        for riqi_excel in ws2['A']:
            riqi_excel_list.append(riqi_excel.value)
        i = 0
        for riqi in riqi_list_1:
            if riqi not in riqi_excel_list:
                ws2.insert_rows(2, 1)
                j = 0
                for dict_name_1, dict_list_1 in data_dict_1.items():
                    ws2.cell(2, 1, riqi)
                    ws2.cell(2, j + 2, dict_list_1[i])
                    j = j + 1
            i = i + 1
        browser.switch_to.default_content()  # 切回到原始界面
        # wb.save('Libresse天猫旗舰店日报表.xlsx')
        wb.save(file_to)

    # Sheet--月汇总
    def sheet_yhz():
        browser.find_element_by_xpath(
            '/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[3]/div[1]/span[2]/span/span').click()
        time.sleep(10)  # 等待数据加载10s
        browser.switch_to.frame('refreshAlarm')
        WebDriverWait(browser, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[17]/transform/div/div[3]/div/visual-modern/div/button')))
        time.sleep(30)
        browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[17]/transform/div/div[3]/div/visual-modern/div/button').click()
        time.sleep(5)
        data_dict_yhz1 = {
            'A21': [],
            'M21': [],
            'F21': [],
            'J21': [],
            'D20': [],
            'N20': [],
            'O20': [],
            'S20': [],
        }
        i = 1
        for data_name_yhz1, data_list_yhz1 in data_dict_yhz1.items():
            data_all_yhz1 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[1]/div[{i + 2}]/div')
            for data_num in data_all_yhz1:
                a = data_num.text
                data_dict_yhz1[data_name_yhz1].append(a)
            data_all_yhz1 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[3]/div[{i + 2}]/div')
            for data_num in data_all_yhz1:
                a = data_num.text
                data_dict_yhz1[data_name_yhz1].append(a)
            i = i + 1
        data_dict_yhz2 = {
            'A20': [],
            'J201': [],
            'J202': [],
            'M20': [],
        }
        i = 1
        for data_name_yhz2, data_list_yhz2 in data_dict_yhz2.items():
            data_all_yhz2 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[2]/div[{i}]/div')
            for data_num in data_all_yhz2:
                b = data_num.text
                data_dict_yhz2[data_name_yhz2].append(b)
            data_all_yhz2 = browser.find_elements_by_xpath(
                f'//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[5]/transform/div/div[3]/div/visual-modern/div/div/div[2]/div[1]/div[4]/div/div[4]/div[{i}]/div')
            for data_num in data_all_yhz2:
                b = data_num.text
                data_dict_yhz2[data_name_yhz2].append(b)
            i = i + 1
        # 数据写入月汇总
        file_to = r'C:\Python\Libresse天猫旗舰店日报表.xlsx'
        # wb = load_workbook('Libresse天猫旗舰店日报表.xlsx')
        wb = load_workbook(file_to)
        ws3 = wb['月汇总']
        ws3['A1'] = this_year + '年'
        data_dict_yhz1.update(data_dict_yhz2)
        i = 0
        for data_name_rhz, data_list_rhz in data_dict_yhz1.items():
            for j in range(len(data_list_rhz)):
                ws3.cell(j + 3, i + 2, data_list_rhz[j])
            i = i + 1
        browser.switch_to.default_content()  # 切回到原始界面
        # wb.save('Libresse天猫旗舰店日报表.xlsx')
        wb.save(file_to)

    # Sheet---总览
    def sheet_zl():
        browser.find_element_by_xpath(
            '/html/body/div/div[1]/div/ul/li[2]/ul/li/ul/div/div[6]/div[2]/div[1]/div[1]/span[2]/span/span').click()
        time.sleep(10)  # 等待数据加载10s
        browser.switch_to.frame('refreshAlarm')
        time.sleep(1)
        WebDriverWait(browser, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[90]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][1]/*[name()="text"]/*[name()="tspan"]')))
        time.sleep(30)
        # 今日时间
        jinrishijian = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[90]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][1]/*[name()="text"]/*[name()="tspan"]').text
        jinrishijian = jinrishijian.replace('-', '/')
        # 今日销售额
        jinrixiaoshou = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[1]/transform/div/div[3]/div/visual-modern/div').text
        # 本月销售额
        benyuexiaoshou = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[4]/transform/div/div[3]/div/visual-modern/div').text
        # 本月目标额
        benyuemubiao = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[41]/transform/div/div[3]/div').text
        # 总达成
        zongdacheng = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[72]/transform/div/div[3]/div/visual-modern/div').text
        # 时间已过
        shijianyiguo = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[71]/transform/div/div[3]/div/visual-modern/div').text.split(
            '过')[1].strip("）")
        # 上周日均
        shangzhourijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[45]/transform/div/div[3]/div/visual-modern/div').text
        # 本周累计销售
        benzhouleijixiaoshou = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[44]/transform/div/div[3]/div/visual-modern/div').text
        # 上周累计销售
        shangzhouleijixiaoshou = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[45]/transform/div/div[3]/div/visual-modern/div').text
        # 上月同期累计
        shangyuetongqileiji = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[82]/transform/div/div[3]/div/visual-modern/div').text
        # 缺口
        quekou = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[69]/transform/div/div[3]/div').text
        # 浏览量
        liulanliang = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[17]/transform/div/div[3]/div/visual-modern/div').text
        liulanliang_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[18]/transform/div/div[3]/div/visual-modern/div').text
        liulanliang_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[46]/transform/div/div[3]/div/visual-modern/div').text
        # 总订单数
        zongdingdanshu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[35]/transform/div/div[3]/div/visual-modern/div').text
        zongdingdanshu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[36]/transform/div/div[3]/div/visual-modern/div').text
        zongdingdanshu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[47]/transform/div/div[3]/div/visual-modern/div').text
        # 总访客数
        zongfangkeshu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[13]/transform/div/div[3]/div/visual-modern/div').text
        zongfangkeshu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[14]/transform/div/div[3]/div/visual-modern/div').text
        zongfangkeshu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[55]/transform/div/div[3]/div/visual-modern/div').text
        # 老访客数
        laofangkeshu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[31]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laofangkeshu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', laofangkeshu.replace(',', '')).group()
        laofangkeshu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[15]/transform/div/div[3]/div/visual-modern/div').text
        laofangkeshu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[57]/transform/div/div[3]/div/visual-modern/div').text
        # 新访客数
        xinfangkeshu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[31]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinfangkeshu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', xinfangkeshu.replace(',', '')).group()
        xinfangkeshu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[16]/transform/div/div[3]/div/visual-modern/div').text
        xinfangkeshu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[56]/transform/div/div[3]/div/visual-modern/div').text
        # 总买家数
        zongmaijiashu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[19]/transform/div/div[3]/div/visual-modern/div').text
        zongmaijiashu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[20]/transform/div/div[3]/div/visual-modern/div').text
        zongmaijiashu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[58]/transform/div/div[3]/div/visual-modern/div').text
        # 老买家数
        laomaijiashu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[32]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laomaijiashu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', laomaijiashu.replace(',', '')).group()
        laomaijiashu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[21]/transform/div/div[3]/div/visual-modern/div').text
        laomaijiashu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[60]/transform/div/div[3]/div/visual-modern/div').text
        # 新买家数
        xinmaijiashu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[32]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinmaijiashu = re.search('([1-9]\d*\.?\d*)|(0\.\d*[1-9])', xinmaijiashu.replace(',', '')).group()
        xinmaijiashu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[22]/transform/div/div[3]/div/visual-modern/div').text
        xinmaijiashu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[59]/transform/div/div[3]/div/visual-modern/div').text
        # 下单率
        xiadanlv = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[33]/transform/div/div[3]/div/visual-modern/div').text
        xiadanlv_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[34]/transform/div/div[3]/div/visual-modern/div').text
        # 老访客下单率
        laoxiadanlv = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[39]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laoxiadanlv_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[51]/transform/div/div[3]/div/visual-modern/div').text
        laoxiadanlv_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[52]/transform/div/div[3]/div/visual-modern/div').text
        # 新访客下单率
        xinxiadanlv = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[39]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinxiadanlv_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[53]/transform/div/div[3]/div/visual-modern/div').text
        xinxiadanlv_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[54]/transform/div/div[3]/div/visual-modern/div').text
        # 客单价
        kedanjia = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[23]/transform/div/div[3]/div/visual-modern/div').text
        kedanjia_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[24]/transform/div/div[3]/div/visual-modern/div').text
        kedanjia_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[91]/transform/div/div[3]/div/visual-modern/div').text
        # 老客单价
        laokedanjia = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[40]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][1]').text
        laokedanjia_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[37]/transform/div/div[3]/div/visual-modern/div').text
        laokedanjia_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[62]/transform/div/div[3]/div/visual-modern/div').text
        # 新客单价
        xinkedanjia = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[40]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"][1]/*[name()="svg"]/*[name()="g"][3]/*[name()="text"][2]').text
        xinkedanjia_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[38]/transform/div/div[3]/div/visual-modern/div').text
        xinkedanjia_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[61]/transform/div/div[3]/div/visual-modern/div').text
        # 主动评价数
        zhudongpingjiashu = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[25]/transform/div/div[3]/div/visual-modern/div').text
        zhudongpingjiashu_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[26]/transform/div/div[3]/div/visual-modern/div').text
        zhudongpingjiashu_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[65]/transform/div/div[3]/div/visual-modern/div').text
        # 正面评价
        zhengmianpingjia = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[30]/transform/div/div[3]/div/visual-modern/div').text
        zhengmianpingjia_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[29]/transform/div/div[3]/div/visual-modern/div').text
        zhengmianpingjia_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[64]/transform/div/div[3]/div/visual-modern/div').text
        # 负面评价
        fumianpingjia = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[28]/transform/div/div[3]/div/visual-modern/div').text
        fumianpingjia_rijun = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[27]/transform/div/div[3]/div/visual-modern/div').text
        fumianpingjia_hb = browser.find_element_by_xpath(
            '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[63]/transform/div/div[3]/div/visual-modern/div').text
        # 淘宝群数据
        data_dict_tb = {
            '活跃人数': [],
            '今日引导成交金额': [],
            '进店购买转化率': [],
            '进店浏览转化率': [],
            '引导成交人数': [],
            '引导进店人数': [],
            '引导客单价': [],
        }
        data_all_tb = browser.find_element_by_class_name('bodyCells')  # Xpath路径太长
        i = 1
        for data_name_tb, data_list_tb in data_dict_tb.items():
            tb1 = data_all_tb.find_element_by_xpath(f'./div/div/div[2]/div[{i}]').text
            tb2 = data_all_tb.find_element_by_xpath(f'./div/div/div[3]/div[{i}]').text
            tb3 = data_all_tb.find_element_by_xpath(f'./div/div/div[4]/div[{i}]').text
            tb4 = data_all_tb.find_element_by_xpath(f'./div/div/div[5]/div[{i}]').text
            data_list_tb.append(tb1)
            data_list_tb.append(tb2)
            data_list_tb.append(tb3)
            data_list_tb.append(tb4)
            i = i + 1
        # 打开Excel模板
        file_to = r'C:\Python\Libresse天猫旗舰店日报表.xlsx'
        # wb = load_workbook('Libresse天猫旗舰店日报表.xlsx')
        wb = load_workbook(file_to)
        ws1 = wb['总览']
        # 数据写入'汇总'
        benyuedijitian = int(jinrishijian.split('/')[2])
        ws1['A1'] = jinrishijian  # 界面拿数据
        ws1['C2'] = benyuexiaoshou
        ws1['C3'] = jinrixiaoshou
        ws1['E2'] = benyuemubiao
        ws1['E3'] = round(int(benyuexiaoshou.replace(',', '')) / benyuedijitian)
        ws1['G2'] = zongdacheng
        ws1['I2'] = shijianyiguo
        ws1['I3'] = quekou

        ws1['D4'] = benzhouleijixiaoshou
        ws1['F4'] = shangzhouleijixiaoshou
        ws1['H4'] = shangyuetongqileiji
        ws1['D5'] = liulanliang
        ws1['D6'] = zongdingdanshu
        ws1['D7'] = zongfangkeshu
        ws1['D8'] = laofangkeshu
        ws1['D9'] = xinfangkeshu
        ws1['D10'] = zongmaijiashu
        ws1['D11'] = laomaijiashu
        ws1['D12'] = xinmaijiashu
        ws1['D13'] = xiadanlv
        ws1['D14'] = laoxiadanlv
        ws1['D15'] = xinxiadanlv
        ws1['D16'] = kedanjia
        ws1['D17'] = laokedanjia
        ws1['D18'] = xinkedanjia
        ws1['D19'] = zhudongpingjiashu
        ws1['D20'] = zhengmianpingjia
        ws1['D21'] = fumianpingjia

        ws1['H5'] = liulanliang_rijun
        ws1['H6'] = zongdingdanshu_rijun
        ws1['H7'] = zongfangkeshu_rijun
        ws1['H8'] = laofangkeshu_rijun
        ws1['H9'] = xinfangkeshu_rijun
        ws1['H10'] = zongmaijiashu_rijun
        ws1['H11'] = laomaijiashu_rijun
        ws1['H12'] = xinmaijiashu_rijun
        ws1['H13'] = xiadanlv_rijun
        ws1['H14'] = laoxiadanlv_rijun
        ws1['H15'] = xinxiadanlv_rijun
        ws1['H16'] = kedanjia_rijun
        ws1['H17'] = laokedanjia_rijun
        ws1['H18'] = xinkedanjia_rijun
        ws1['H19'] = zhudongpingjiashu_rijun
        ws1['H20'] = zhengmianpingjia_rijun
        ws1['H21'] = fumianpingjia_rijun

        ws1['J5'] = liulanliang_hb
        ws1['J6'] = zongdingdanshu_hb
        ws1['J7'] = zongfangkeshu_hb
        ws1['J8'] = laofangkeshu_hb
        ws1['J9'] = xinfangkeshu_hb
        ws1['J10'] = zongmaijiashu_hb
        ws1['J11'] = laomaijiashu_hb
        ws1['J12'] = xinmaijiashu_hb
        ws1['J14'] = laoxiadanlv_hb
        ws1['J15'] = xinxiadanlv_hb
        ws1['J16'] = kedanjia_hb
        ws1['J17'] = laokedanjia_hb
        ws1['J18'] = xinkedanjia_hb
        ws1['J19'] = zhudongpingjiashu_hb
        ws1['J20'] = zhengmianpingjia_hb
        ws1['J21'] = fumianpingjia_hb

        i = 0
        for data_name_tb, data_list_tb in data_dict_tb.items():
            ws1[f'D{i + 22}'] = data_list_tb[0]
            ws1[f'F{i + 22}'] = data_list_tb[1]
            ws1[f'H{i + 22}'] = data_list_tb[2]
            ws1[f'J{i + 22}'] = data_list_tb[3]
            i = i + 1
        browser.switch_to.default_content()  # 切回到原始界面
        # wb.save('Libresse天猫旗舰店日报表.xlsx')
        wb.save(file_to)
        wb.close()

    sheet_rdp()
    sheet_rhz()
    sheet_yhz()
    sheet_zl()
    browser.quit()

# 获取Excel数据
get_excel()
# 日期时间
today = datetime.date.today()
oneday = datetime.timedelta(days=1)
yesterday = str(today - oneday)
nowtime = time.strftime("%Y-%m-%d", time.localtime(time.time()))
nowtime2 = time.strftime("%H:%M", time.localtime(time.time()))
yesterday_list = str(today - oneday).replace('-0', '-').split('-')
yesterday_2 = (yesterday_list[0] + '-' + yesterday_list[1] + '-' + yesterday_list[2])
# 定义浏览器
option = webdriver.ChromeOptions()
option.add_argument("--headless")
browser = webdriver.Chrome(options=option)

# 图片加水印
def add_text_to_image(pic_file, text):
    image = Image.open(pic_file)
    font = ImageFont.truetype(r'C:\Windows\Fonts\msyh.ttc', 40)
    new_img = Image.new('RGBA', (image.size[0] * 3, image.size[1] * 3), (0, 0, 0, 0))
    new_img.paste(image, image.size)
    font_len = len(text)
    rgba_image = new_img.convert('RGBA')
    text_overlay = Image.new('RGBA', rgba_image.size, (255, 255, 255, 0))
    image_draw = ImageDraw.Draw(text_overlay)
    for i in range(0, rgba_image.size[0], font_len*80+360):
        for j in range(0, rgba_image.size[1], 200):
            image_draw.text((i, j), text, font=font, fill=(0, 0, 0, 20))
    text_overlay = text_overlay.rotate(30)      # 调整水印方向
    image_with_text = Image.alpha_composite(rgba_image, text_overlay)
    image_with_text = image_with_text.crop((image.size[0], image.size[1], image.size[0] * 2, image.size[1] * 2))
    image_with_text.save(pic_file)

# 获取截图
def portallogin():
    total_num = 0
    browser.set_window_size(1920, 1080)
    time.sleep(2)
    name_dict = {
        '数据部': "数据部",
        # '133437':"京东业务部",
        # '133440': "天猫超市业务部",
        # '133444': "旗舰店业务部",
        # '133445': "成护品类发展部",
        # '148608': "女护品类发展部",
        # '140088': "淘系业务部",
        # '146995':"综合业务一部",
        # '139739': "综合业务二部",
        # '137576': "拼多多业务部",
        # '148505': "微商城",
        # '108043': "推广一部",
        # '138996': "推广三部",
        # '136153': "投放部"
        # '148896': "投放部",
        # '146549': "推广部",
    }
    for user, user_name in name_dict.items():
        total_num = 0
        dirs_a, file_name_a = os.path.split(__file__)
        pic_zl = os.path.join(dirs_a, f"Libresse官方旗舰店日报表/{user}总览.png")
        pic_zla2 = os.path.join(dirs_a, f"Libresse官方旗舰店日报表/{user}总览2.png")
        if user == "数据部":
            url_zla = "http://100.100.0.5:8084/DemoDL.html?UserName=bidev&ReportId=a3bc1940-ed0b-48a0-9987-d988f2475762|9e1cfb0e-5c6d-47a7-852e-e796836a499a&YZ=false"
            url_zla2 = "http://100.100.0.5:8084/DemoDL.html?UserName=bidev&ReportId=f07db9e2-20b4-49d7-a290-b7010076585b|9e1cfb0e-5c6d-47a7-852e-e796836a499a&YZ=false"
            # 第一张图片
            browser.get(url_zla)
            WebDriverWait(browser, 60, 0.5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="embedContainer"]/iframe')))
            time.sleep(30)
            # 判断日期
            frame_xpath = browser.find_element_by_xpath('//*[@id="embedContainer"]/iframe')
            browser.switch_to.frame(frame_xpath)
            time.sleep(1)
            riqi1 = browser.find_element_by_xpath(
                '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas-modern/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container-modern[79]/transform/div/div[3]/div/visual-modern/div/*[name()="svg"]/*[name()="g"][1]/*[name()="text"]/*[name()="tspan"]').text
            # 判断图片大小
            browser.save_screenshot(pic_zl)
            print(f'获取{user}总览1成功')
            # 第二张图片
            browser.get(url_zla2)
            WebDriverWait(browser, 60, 0.5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="embedContainer"]/iframe')))
            time.sleep(30)
            browser.save_screenshot(pic_zla2)
            print(f'获取{user}总览2成功')
            time.sleep(1)
            # 新增第二部分截图
            im = Image.open(pic_zla2)
            im = im.crop((0, 40, 1920, 428))
            im.save(pic_zla2)
            # 打开图片加水印
            add_text_to_image(pic_zl, 'Libresse天猫旗舰店')
            add_text_to_image(pic_zla2, 'Libresse天猫旗舰店')
            browser.set_window_size(1920, 1380)
            time.sleep(0.5)
            browser.set_window_size(1920, 1080)
            pic1_size = round(os.path.getsize(pic_zl) / 1024, 2)
            pic2_size = round(os.path.getsize(pic_zla2) / 1024, 2)
            if riqi1 == yesterday_2 and pic1_size > 360 and pic2_size > 40:
                total_num = 0
            elif riqi1 != yesterday_2:
                total_num = 1
                break
            elif pic1_size < 360 or pic2_size < 40:
                total_num = 2
                break
            else:
                total_num = 3
                break
        else:
            url_zl = f"http://100.100.0.5:8084/DemoDL.html?UserName={user}&ReportId=a3bc1940-ed0b-48a0-9987-d988f2475762|9e1cfb0e-5c6d-47a7-852e-e796836a499a&YZ=true"
            url_zl2 = f"http://100.100.0.5:8084/DemoDL.html?UserName={user}&ReportId=f07db9e2-20b4-49d7-a290-b7010076585b|9e1cfb0e-5c6d-47a7-852e-e796836a499a&YZ=true"
            browser.get(url_zl)
            # 返回总览 等待20S截图
            time.sleep(60)
            browser.save_screenshot(pic_zl)
            browser.get(url_zl2)
            time.sleep(60)
            browser.save_screenshot(pic_zla2)
            time.sleep(1)
            im = Image.open(pic_zla2)
            im = im.crop((0, 40, 1920, 428))
            im.save(pic_zla2)
            # 打开图片加水印
            add_text_to_image(pic_zl, 'Libresse天猫旗舰店')
            add_text_to_image(pic_zla2, 'Libresse天猫旗舰店')
            browser.set_window_size(1920, 1380)
            time.sleep(0.5)
            browser.set_window_size(1920, 1080)
    return total_num

# 定义发邮件函数
def sendemail():
    send_dict = {
        '数据部': ['lin.wh@vinda.com', 'f.yj@vinda.com', 'xu.ch@vinda.com', 'su.sj@vinda.com', 'xuyuxian@vinda.com',
                'liangyy@vinda.com', 'tan.qing@vinda.com', 'Wu.xiaoyan@vinda.com', 'zhou.pu@vinda.com', 'mo.cp@vinda.com',
                'li.sy@vinda.com', 'guo.juan@vinda.com', 'zhu.sk@vinda.com', 'luo.bj@vinda.com'],
        # '133437': ['hou.yan@vinda.com', 'zheng.yi@vinda.com', 'li.yu@vinda.com','wang.ke@vinda.com',
        #            'tu.jk@vinda.com', 'zhan.ww@vinda.com', 'tang.ml@vinda.com', 'zhou.huan@vinda.com',
        #            'yangxueli@vinda.com', 'tian.ht@vinda.com', 'tang.jy@vinda.com', 'pan.zy@vinda.com'],
        # '133440': ['li.li@vinda.com', 'wangxiao@vinda.com', 'wu.wy@vinda.com', 'zhao.jing@vinda.com', 'l.yi@vinda.com',
        #            'wangyx@vinda.com', 'yuan.jw@vinda.com', 'w.xin@vinda.com', 'zhouyinglan@vinda.com',
        #            'jin.zh@vinda.com',
        #            'liukang@vinda.com', 'chen.bb@vinda.com'],
        # '133444': ['ruan.cj@vinda.com', 'huang.yz@vinda.com', 'chensihuan@vinda.com', 'zhang.jing@vinda.com',
        #            'c.meng@vinda.com', 'yang.wc@vinda.com', 'wu.yp@vinda.com', 'lao.sl@vinda.com', 'huangyz@vinda.com'],
        # '133445': ['zhong.th@vinda.com', 'shan.cc@vinda.com', 'zhu.zy@vinda.com', 'bao.pw@vinda.com', 'xuxp@vinda.com',
        #            'wuyj@vinda.com', 'liang.kun@vinda.com', 'lin.yx@vinda.com','fu.ly@vinda.com','xuzj@vinda.com',
        #            'huangying@vinda.com', 'su.rn@vinda.com'],
        # '148608': ['f.yj@vinda.com', 'xu.ch@vinda.com', 'su.sj@vinda.com', 'xuyuxian@vinda.com', 'zhou.pu@vinda.com',
        #            'mo.cp@vinda.com', 'Wu.xiaoyan@vinda.com', 'tan.qing@vinda.com'],
        # '140088': ['ding.jie@vinda.com', 'c.lj@vinda.com', 'zhang.kf@vinda.com', 'cen.wj@vinda.com'],
        # '146995': ['luo.yang@vinda.com', 'huang.yanling@vinda.com', 'wang.hw@vinda.com', 'linguojuan@vinda.com',
        #            'li.ss@vinda.com', 'li.wz@vinda.com','zengxl@vinda.com','chen.kx@vinda.com','feng.qw@vinda.com'],
        # '139739': ['gan.yl@vinda.com', 'liuzw@vinda.com', 'jianni@vinda.com', 'cheng.dong@vinda.com',
        #            'wang.zx@vinda.com',
        #            'ly.jt@vinda.com', 'ding.ss@vinda.com'],
        # '137576': ['bai.rx@vinda.com', 'zheng.yu@vinda.com', 'zheng.kai@vinda.com', 'tian.zh@vinda.com','luoqing@vinda.com'],
        # '148505': ['lin.gj@vinda.com', 'chenyushi@vinda.com'],
        # '108043': ['tan.qing@vinda.com','ji.hh@vinda.com', 'yuanyuan@vinda.com', 'lijiawen@vinda.com', 'zeng.xl@vinda.com',
        #            'chenyl@vinda.com',
        #            'lin.mian@vinda.com', 'zhang.ws@vinda.com', 'guo.yj@vinda.com', 'liu.mei@vinda.com',
        #            'dugy@vinda.com','luo.zp@vinda.com','tan.pt@vinda.com',
        #            ],
        # '138996': ['tan.qing@vinda.com', 'ou.mx@vinda.com', 'huang.mt@vinda.com', 'mo.mh@vinda.com',
        #            'Wu.xiaoyan@vinda.com', 'xiao.ss@vinda.com', 'luo.ms@vinda.com', 'xiehy@vinda.com',
        #            'xie.hx@vinda.com',
        #            'liang.hj@vinda.com', 'yu.jw@vinda.com'],
        # '136153': ['cai.by@vinda.com', 'zhang.xm@vinda.com', 'Hu.yx@vinda.com', 'lvxh@vinda.com',
        #            'zhanghuanyu@vinda.com',
        #            'feng.hx@vinda.com', 'zhou.pu@vinda.com', 'zhanting@vinda.com', 'lijia@vinda.com', 'yehm@vinda.com',
        #            'chenzy@vinda.com', 'liwm@vinda.com']
        # '148896': ['zhou.pu@vinda.com', 'mo.cp@vinda.com'],
        # '146549': ['Wu.xiaoyan@vinda.com', 'tan.qing@vinda.com'],
    }
    # 报表访问网址
    ScreenshotUrl = 'http://pbi.vinda.com'
    # 设置发送人
    # sender = 'jiqiren@vinda.com'
    # 邮箱服务器
    smtpserver = 'mail.vinda.com'
    username = 'rst_admin@vinda.com'
    password = 'RV@26137'   # 授权码,注意不是邮箱登录密码
    sender = 'rst_admin@vinda.com'
    # 发送邮件
    smtp = smtplib.SMTP()
    smtp.connect(smtpserver, 25)
    smtp.login(username, password)
    # 取图片地址发送邮件
    for ro_name, ro_email in send_dict.items():
        dirs_a, file_name_a = os.path.split(__file__)
        zl_file = os.path.join(dirs_a, f"Libresse官方旗舰店日报表/{ro_name}总览.png")
        zl_file2 = os.path.join(dirs_a, f"Libresse官方旗舰店日报表/{ro_name}总览2.png")
        piclist = [zl_file, zl_file2]
        to_reciver = ro_email
        cc_reciver = ['guo.juan@vinda.com', 'luo.bj@vinda.com', 'maggie.huang@xunyisoft.com',
                      'Jeremy.Zhang@xunyisoft.com', 'Leo.Zhang@xunyisoft.com']
        # to_reciver = ['maggie.huang@xunyisoft.com']
        # cc_reciver = ['Leo.Zhang@xunyisoft.com']
        receiver = to_reciver + cc_reciver
        subject = " 【" + str(today)[-5:]+" "+str(nowtime2) + '】 Libresse官方旗舰店日报表'
        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = sender
        # 收件人为多个收件人,通过join将列表转换为以;为间隔的字符串
        msg['To'] = ";".join(to_reciver)
        msg['Cc'] = ";".join(cc_reciver)
        # alternative方式
        msgAlternative = MIMEMultipart('alternative')
        msg.attach(msgAlternative)
        mail_body = "<p>截止至" + yesterday + "，Libresse官方旗舰店销售概况如下图所示，如需更详细的数据请点击下面链接查看报表。</p>" \
            """
            <p><a href='""" + ScreenshotUrl + """'>总览报表查看详情链接</a></p>
            <p>Libresse天猫官方旗舰店销售概况：</p>
            <p><img src="cid:send_image0" ></p>
            <p><img src="cid:send_image1" ></p>
            """
        msgText = (MIMEText(mail_body, 'html', 'utf-8'))
        msgAlternative.attach(msgText)
        # 新增附件
        file_path = r'C:\Python\Libresse天猫旗舰店日报表.xlsx'
        att1 = MIMEText(open(file_path, 'rb').read(), 'base64', 'utf-8')
        att1["Content-Type"] = 'application/octet-stream'
        att1.add_header('Content-Disposition', 'attachment', filename="Libresse天猫旗舰店日报表.xlsx")
        msg.attach(att1)
        for k, file in enumerate(piclist):
            src_cid = "<send_image" + str(k) + ">"
            # 指定图片为当前目录
            fp = open(file, 'rb')
            msgImage = MIMEImage(fp.read())
            fp.close()
            # 定义图片ID,在HTML文本中引用
            msgImage.add_header('Content-ID', src_cid)
            msg.attach(msgImage)
        smtp.sendmail(sender, receiver, msg.as_string())
        time.sleep(1)
    smtp.quit()
    smtp.close()

# 记录状态
def status_txt(num):
    with open('status_5.txt', 'w', encoding='utf-8') as f:
        f.write(num)

status_txt('4')
status_num = portallogin()
browser.quit()
if status_num == 0:
    sendemail()
    status_txt('0')
    print('邮件发送成功')
if status_num == 1:
    status_txt('1')
    print('日期核对错误')
if status_num == 2:
    status_txt('2')
    print('图片大小错误')
else:
    status_txt('3')
    print('日期和图片错误')
time.sleep(10)
os.system("taskkill /f /t /im chromedriver.exe")
